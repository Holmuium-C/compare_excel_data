
#!/usr/bin/env python
# -*- coding: utf-8 -*-
# 导入模块 openpyxl
import openpyxl
import sys
import re
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

##循环求差异

def get_out_trade_no(wb):
    sheet = wb.sheetnames
    sheet = wb.active
    excel_out_trade_no = []
    for cellobj in sheet['A']:
        excel_out_trade_no.append(cellobj.value)
    return excel_out_trade_no

i = int(input("请输入需要对比表的数量"))
excelname = []
name = ""
for n in range(0, i):
    excelname1 = input(
        "请输入需要对比的表，相对路径直接输入./+表名（每次输入一个excel）")
    a = excelname1
    excelname.append(a)
for n in range(0, i):

    locals()[str(n)+"_wb"] = openpyxl.load_workbook(excelname[n])
    locals()[str(n)+"get_out_trade_no"] = get_out_trade_no(locals()[str(n)+"_wb"])

    locals()[str(n)+"S"] = set(locals()[str(n)+"get_out_trade_no"])

for n in range(0,i-1,1):

    if i==2:
        difference = list(locals()[str(0)+"S"] ^ locals()[str(1)+"S"])
        print("=======================存在差异的数据如下===2==========================\n")
        print(excelname[0]+"表存在而"+excelname[1]+"表不存在的差异数据")
        for x in locals()[str(1)+"S"].intersection(difference):
            print(x)
            
            sheet0 = locals()[str(0)+"_wb"].active
            sheet0Value = []
            for cellobj in sheet0['A']:
                if cellobj.value in difference:
                    cellobj.font = Font(color=colors.BLACK,italic=True,bold=True)
                    cellobj.fill = PatternFill("solid",fgColor="FF0000")
                    locals()[str(0)+"xlsx"] = str(excelname[0])
                    xlsxname = re.findall(r"./(.+?).xlsx" , locals()[str(0)+"xlsx"])
                    locals()[str(0)+"_wb"].save("./"+str(xlsxname)+"_Difference.xlsx")
                    

        print(excelname[1]+"表存在而"+excelname[0]+"表不存在的差异数据")
        for y in locals()[str(0)+"S"].intersection(difference):
            print(y)

            sheet1 = locals()[str(1)+"_wb"].active
            sheet1Value = []
            for cellobj in sheet1['A']:
                if cellobj.value in difference:
                    cellobj.font = Font(color=colors.BLACK,italic=True,bold=True)
                    cellobj.fill = PatternFill("solid",fgColor="FF0000")
                    locals()[str(1)+"xlsx"] = str(excelname[1])
                    xlsxname = re.findall(r"./(.+?).xlsx" , locals()[str(1)+"xlsx"])
                    locals()[str(1)+"_wb"].save("./"+str(xlsxname)+"_Difference.xlsx")
        break
    
    if i>2:
        a=n+1
        for j in range(a,i,1):
            print(locals()[str(n)+"S"])
            print(locals()[str(j)+"S"])
            difference = list(locals()[str(n)+"S"] ^ locals()[str(j)+"S"])
            print("=======================存在差异的数据如下==>2==========================\n")
            print(excelname[n]+"（A）表存在而"+excelname[j]+"(B)表不存在的差异数据")
            for x in locals()[str(n)+"S"].intersection(difference):
                print(x)

                locals()["sheet"+str(n)] = locals()[str(n)+"_wb"].active
                for cellobj in locals()["sheet"+str(n)]['A']:
                    if cellobj.value in difference:
                        cellobj.font = Font(color=colors.BLACK,italic=True,bold=True)
                        cellobj.fill = PatternFill("solid",fgColor="FF0000")
                        locals()[str(n)+"xlsx"] = str(excelname[n])
                        xlsxname = re.findall(r"./(.+?).xlsx" , locals()[str(n)+"xlsx"])
                        # print(str(xlsxname)+"name")
                        locals()[str(n)+"_wb"].save("./"+str(xlsxname)+"_Difference.xlsx")


            print(excelname[j]+"(B)表存在而"+excelname[n]+"(A)表不存在的差异数据")
            for y in locals()[str(j)+"S"].intersection(difference):
                print(y)
                locals()["sheet"+str(j)] = locals()[str(j)+"_wb"].active
                for cellobj in locals()["sheet"+str(j)]['A']:
                    if cellobj.value in difference:
                        cellobj.font = Font(color=colors.BLACK,italic=True,bold=True)
                        cellobj.fill = PatternFill("solid",fgColor="FF0000")
                        locals()[str(j)+"xlsx"] = str(excelname[j])
                        xlsxname2 = re.findall(r"./(.+?).xlsx" , locals()[str(j)+"xlsx"])
                        # print(str(xlsxname))
                        locals()[str(j)+"_wb"].save("./"+str(xlsxname2)+"_Difference.xlsx")


print("差异excel文件已保存至相对路径")

ip = input("\n\n\n=========输入0退出=========\n")
if ip  == "0":
    sys.exit(0)

